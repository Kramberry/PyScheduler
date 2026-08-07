from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import LETTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import json
import os
import sys
import shutil
import webbrowser
import logging
from datetime import datetime, timedelta
APP_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.abspath(".")
EMP_FILE = os.path.join(APP_DIR, "employees.json")
SCHEDULE_FILE = os.path.join(APP_DIR, "last_schedule.json")
ROLES_FILE = os.path.join(APP_DIR, "roles.json")
HISTORY_FILE = os.path.join(APP_DIR, "schedule_history.json")
EMP_BACKUP_FILE = os.path.join(APP_DIR, "employees_backup.json")
SCHEDULE_BACKUP_FILE = os.path.join(APP_DIR, "last_schedule_backup.json")
HISTORY_BACKUP_FILE = os.path.join(APP_DIR, "schedule_history_backup.json")

logging.basicConfig(filename=os.path.join(APP_DIR, "app.log"), level=logging.INFO)

def resource_path(relative: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base, relative)

app = Flask(
    __name__,
    template_folder=resource_path("templates"),
)
app.secret_key = os.environ.get("SHIFTDESK_SECRET_KEY", "shiftdesk-local-dev-key")

def to_24h(time_str):
    if not time_str:
        return ''
    try:
        return datetime.strptime(time_str.strip(), "%I:%M %p").strftime("%H:%M")
    except Exception:
        return ''

app.jinja_env.filters['to24h'] = to_24h


DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


# ------------------------------------
# Helpers to load/save employees
# ------------------------------------
def load_employees():
    if not os.path.exists(EMP_FILE):
        save_employees([])
        return []

    with open(EMP_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_employees(employees_list):
    with open(EMP_FILE, "w", encoding="utf-8") as f:
        json.dump(employees_list, f, indent=2)

def load_last_schedule():
    if not os.path.exists(SCHEDULE_FILE):
        return {}
    with open(SCHEDULE_FILE, "r", encoding="utf-8") as f:
        return json.load(f) 

def save_last_schedule(data):
    with open(SCHEDULE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


# ------------------------------------
# Shift roles — canonical list, separate from any one person or shift.
# Only touched at deliberate save points (schedule save/export/print), never
# on every keystroke in the role-input box.
# ------------------------------------
def load_roles():
    if not os.path.exists(ROLES_FILE):
        seeded = sorted(extract_roles_from_schedule(load_last_schedule().get("schedule", {})))
        save_roles(seeded)
        return seeded
    with open(ROLES_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_roles(roles_list):
    with open(ROLES_FILE, "w", encoding="utf-8") as f:
        json.dump(sorted(set(roles_list)), f, indent=2)


def extract_roles_from_schedule(schedule_dict):
    roles = set()
    for day_data in schedule_dict.values():
        for cell in day_data.values():
            for r in cell.get("role", []):
                if r and r != "PTO":
                    roles.add(r)
    return roles


def register_roles(schedule_dict):
    used = extract_roles_from_schedule(schedule_dict)
    known = set(load_roles())
    if used - known:
        save_roles(known | used)


# ------------------------------------
# Multi-week history — each save archives that week's schedule keyed by its
# own week_start, so other weeks are never touched.
# ------------------------------------
def load_history():
    if not os.path.exists(HISTORY_FILE):
        seeded = {}
        last = load_last_schedule()
        if last.get("week_start") and last.get("schedule"):
            seeded[last["week_start"]] = last["schedule"]
        save_history(seeded)
        return seeded
    with open(HISTORY_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_history(history_dict):
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history_dict, f, indent=2)


def archive_week(week_start, schedule_dict):
    if not week_start:
        return
    history = load_history()
    history[week_start] = schedule_dict
    save_history(history)


# ------------------------------------
# Pre-delete backup, so removing an employee can be undone
# ------------------------------------
def backup_employee_data():
    if os.path.exists(EMP_FILE):
        shutil.copyfile(EMP_FILE, EMP_BACKUP_FILE)
    if os.path.exists(SCHEDULE_FILE):
        shutil.copyfile(SCHEDULE_FILE, SCHEDULE_BACKUP_FILE)
    if os.path.exists(HISTORY_FILE):
        shutil.copyfile(HISTORY_FILE, HISTORY_BACKUP_FILE)


def restore_employee_backup():
    restored = False
    if os.path.exists(EMP_BACKUP_FILE):
        shutil.copyfile(EMP_BACKUP_FILE, EMP_FILE)
        restored = True
    if os.path.exists(SCHEDULE_BACKUP_FILE):
        shutil.copyfile(SCHEDULE_BACKUP_FILE, SCHEDULE_FILE)
        restored = True
    if os.path.exists(HISTORY_BACKUP_FILE):
        shutil.copyfile(HISTORY_BACKUP_FILE, HISTORY_FILE)
        restored = True
    return restored


def remove_employees_from_schedules(names):
    """Surgically drop the given employee names from the current schedule and
    every archived week — everyone else's entries are left untouched."""
    names = set(names)
    if not names:
        return

    last = load_last_schedule()
    if last.get("schedule"):
        for name in names:
            last["schedule"].pop(name, None)
        save_last_schedule(last)

    history = load_history()
    changed = False
    for week_schedule in history.values():
        for name in names:
            if week_schedule.pop(name, None) is not None:
                changed = True
    if changed:
        save_history(history)


# ------------------------------------
# Calculate hours from "9:00 - 6:00"
# ------------------------------------
def calculate_hours(cell_value: str) -> float:
    if not cell_value:
        return 0
    if "PTO" in cell_value:
        return 8
    
    try:
        time_line = cell_value.split("\n")[0]  # "09:00 - 6:00pm"
        start_str, end_str = [t.strip() for t in time_line.split("-")]

        def to_minutes_12hr(t):
            time_part, period = t.split()
            hour, minute = map(int, time_part.split(":"))
            if period == "PM" and hour != 12:
                hour += 12
            if period == "AM" and hour == 12:
                hour = 0

            return hour * 60 + minute
        start = to_minutes_12hr(start_str)
        end = to_minutes_12hr(end_str)

        if end < start:
            return 0  # prevents overnight shifts for now
        
        return (end - start) / 60.0
    except Exception:
        return 0


# ------------------------------------
# Shared week-data shaping, used by the single-week and multi-week
# Excel/PDF exports so the two formats can never drift apart.
# ------------------------------------
def build_week_rows(employees, schedule_dict):
    """Returns [(employee, per_day, total_hours), ...] for employees who have
    at least one day of data this week. per_day[i] describes DAYS[i] as
    {'roles': [...], 'start': str, 'end': str, 'pto': bool}."""
    rows = []
    for emp in employees:
        day_data = schedule_dict.get(emp, {}) if schedule_dict else {}
        per_day = []
        total = 0.0
        has_data = False
        for day in DAYS:
            cell = day_data.get(day, {}) if day_data else {}
            roles = [r for r in (cell.get("role") or []) if r and r.strip()]
            start = (cell.get("start") or "").strip()
            end = (cell.get("end") or "").strip()
            is_pto = "PTO" in roles
            if is_pto:
                total += 8
                has_data = True
            elif start and end:
                total += calculate_hours(f"{start} - {end}")
                has_data = True
            elif roles:
                has_data = True
            per_day.append({"roles": roles, "start": start, "end": end, "pto": is_pto})
        if has_data:
            rows.append((emp, per_day, total))
    return rows


def cell_excel_text(day):
    if day["pto"]:
        return "PTO"
    role_text = ", ".join(r for r in day["roles"] if r != "PTO")
    if day["start"] and day["end"]:
        return f"{day['start']} - {day['end']}\n{role_text}" if role_text else f"{day['start']} - {day['end']}"
    return role_text


DAY_COLORS = {
    2: {"header": "2563EB", "light": "EFF6FF", "band": "DBEAFE"},  # Monday    - blue
    3: {"header": "7C3AED", "light": "F5F3FF", "band": "EDE9FE"},  # Tuesday   - violet
    4: {"header": "0D9488", "light": "F0FDFA", "band": "CCFBF1"},  # Wednesday - teal
    5: {"header": "EA580C", "light": "FFF7ED", "band": "FFEDD5"},  # Thursday  - orange
    6: {"header": "DB2777", "light": "FDF2F8", "band": "FCE7F3"},  # Friday    - pink
}
TOTAL_HEADER_COLOR = "16A34A"
TOTAL_BG_COLOR = "DCFCE7"
TOTAL_FONT_COLOR = "15803D"


def write_week_sheet(ws, date_range_text, rows):
    """Fills a freshly created worksheet with one week's schedule, styled to
    match the app's export look. Used for both single- and multi-week Excel."""
    headers = ["Team Member"] + DAYS + ["Total Hours"]
    num_cols = len(headers)

    for col_idx, text in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=text)
    for row_idx, (emp, per_day, total) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=emp)
        for col_idx, day in enumerate(per_day, start=2):
            ws.cell(row=row_idx, column=col_idx, value=cell_excel_text(day))
        ws.cell(row=row_idx, column=num_cols, value=f"{total:.1f}")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    ws.column_dimensions["A"].width = 18
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 26
    ws.column_dimensions["G"].width = 14

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Weekly Schedule  ·  {date_range_text}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E3A5F", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[2], start=1):
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in DAY_COLORS:
            cell.fill = PatternFill(start_color=DAY_COLORS[col_idx]["header"], fill_type="solid")
        elif col_idx == num_cols:
            cell.fill = PatternFill(start_color=TOTAL_HEADER_COLOR, fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="1E3A5F", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for band_idx, row in enumerate(ws.iter_rows(min_row=3)):
        is_band = band_idx % 2 == 1
        for col_idx, cell in enumerate(row, start=1):
            val = str(cell.value or "").strip()
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

            if col_idx == 1:  # Team Member
                cell.font = Font(bold=True, size=10)
                if is_band:
                    cell.fill = PatternFill(start_color="F8FAFC", fill_type="solid")
            elif col_idx == num_cols:  # Total Hours
                cell.font = Font(bold=True, size=10, color=TOTAL_FONT_COLOR)
                cell.fill = PatternFill(start_color=TOTAL_BG_COLOR, fill_type="solid")
            elif col_idx in DAY_COLORS:  # Day columns
                dc = DAY_COLORS[col_idx]
                if val == "PTO":
                    cell.font = Font(bold=True, size=10, color="92400E")
                    cell.fill = PatternFill(start_color="FEF3C7", fill_type="solid")
                elif not val:
                    cell.font = Font(size=10)
                    cell.fill = PatternFill(start_color="F1F5F9", fill_type="solid")
                else:
                    cell.font = Font(size=10)
                    cell.fill = PatternFill(start_color=dc["band"] if is_band else dc["light"], fill_type="solid")

    ws.freeze_panes = "A3"


def safe_sheet_title(existing_titles, start_date):
    """Excel tab names: <=31 chars, unique. Dates alone satisfy both, but we
    guard duplicates defensively rather than assume it."""
    base = start_date.strftime("%b %d, %Y")[:31]
    title = base
    n = 2
    while title in existing_titles:
        suffix = f" ({n})"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    return title


def week_date_range_text(week_start):
    start_date = datetime.strptime(week_start, "%Y-%m-%d")
    end_date = start_date + timedelta(days=4)
    return start_date, end_date, f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"


def cell_pdf_paragraph(day, cell_style, pto_style):
    if day["pto"]:
        return Paragraph("PTO", pto_style)
    role_text = ", ".join(r for r in day["roles"] if r != "PTO")
    if day["start"] and day["end"]:
        text = f"{day['start']} - {day['end']}"
        if role_text:
            text += f"<br/>{role_text}"
        return Paragraph(text, cell_style)
    if role_text:
        return Paragraph(role_text, cell_style)
    return ""


def build_pdf_week_elements(date_range_text, rows):
    styles = getSampleStyleSheet()
    elements = [Paragraph("<b>Weekly Schedule</b>", styles["Title"])]
    if date_range_text:
        elements.append(Paragraph(date_range_text, styles["Normal"]))
    elements.append(Spacer(1, 16))

    cell_style = ParagraphStyle('cell', fontSize=8, leading=11)
    pto_style = ParagraphStyle('pto', fontSize=8, leading=11,
                                textColor=colors.HexColor('#92400E'), fontName='Helvetica-Bold')

    table_data = [["Employee"] + DAYS + ["Total Hours"]]
    for emp, per_day, total in rows:
        row = [emp] + [cell_pdf_paragraph(d, cell_style, pto_style) for d in per_day] + [f"{total:.1f}"]
        table_data.append(row)

    col_widths = [80] + [81] * len(DAYS) + [55]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID",       (0, 0), (-1, -1), 1, colors.black),
        ("BACKGROUND", (0, 0), (-1,  0), colors.lightgrey),
        ("FONT",       (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONT",       (0, 1), ( 0, -1), "Helvetica-Bold"),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
    ]))
    elements.append(table)
    return elements


# ------------------------------------
# MAIN SCHEDULE PAGE
# ------------------------------------
@app.route("/", methods=["GET"])
def index():
    employees = load_employees()
    saved = load_last_schedule()
    return render_template(
        "schedule_form.html",
        employees=employees,
        days=DAYS,
        saved=saved,
        roles=load_roles()
    )


# ------------------------------------
# PRINT PREVIEW
# ------------------------------------
@app.route("/print-preview", methods=["POST"])
def print_preview():
    employees = load_employees()
    week_start = request.form.get("week_start")

    saved_data = {
        "week_start": week_start,
        "schedule": {}
    }
    for emp in employees:
        saved_data["schedule"][emp] = {}
        for day in DAYS:
            saved_data["schedule"][emp][day] = {
                "role": request.form.getlist(f"{emp}_{day}_role"),
                "start": request.form.get(f"{emp}_{day}_start", ""),
                "end": request.form.get(f"{emp}_{day}_end", "")
            }
    save_last_schedule(saved_data)
    register_roles(saved_data["schedule"])
    archive_week(week_start, saved_data["schedule"])

    totals = {}
    for emp in employees:
        total = 0.0
        for day in DAYS:
            cell = saved_data["schedule"][emp][day]
            roles = cell["role"]
            if "PTO" in roles:
                total += 8.0
            elif cell["start"] and cell["end"]:
                total += calculate_hours(f"{cell['start']} - {cell['end']}")
        totals[emp] = total

    role_color_palette = [
        "#2563EB", "#7C3AED", "#0D9488", "#EA580C", "#DB2777",
        "#16A34A", "#D97706", "#DC2626", "#0891B2", "#7C2D12",
    ]
    role_colors = {}
    for emp in employees:
        for day in DAYS:
            for role in saved_data["schedule"][emp][day]["role"]:
                if role and role != "PTO" and role not in role_colors:
                    role_colors[role] = role_color_palette[len(role_colors) % len(role_color_palette)]

    return render_template(
        "print_preview.html",
        employees=employees,
        days=DAYS,
        saved=saved_data,
        totals=totals,
        role_colors=role_colors,
    )


# ------------------------------------
# MULTI-WEEK EXPORT
# ------------------------------------
@app.route("/export-multi", methods=["GET"])
def export_multi_picker():
    history = load_history()
    weeks = []
    for week_start in history.keys():
        try:
            _start, _end, label = week_date_range_text(week_start)
        except ValueError:
            continue
        weeks.append({"week_start": week_start, "label": label})
    weeks.sort(key=lambda w: w["week_start"], reverse=True)
    return render_template("multi_export.html", weeks=weeks)


def resolve_export_weeks(weeks_json_str, week_starts_list):
    """Returns (sorted week_starts, {week_start: schedule_dict}).

    The tabbed editor sends full schedule data for every open tab as
    `weeks_json` — exporting is the deliberate save point, so each of those
    weeks is archived (and its roles registered) as a side effect, the same
    way a single-week export always has been. The standalone "From Archive"
    picker page instead sends `week_starts` referencing weeks already
    archived in an earlier session, so no re-archiving is needed there.
    """
    if weeks_json_str:
        try:
            payload = json.loads(weeks_json_str)
        except (TypeError, ValueError):
            return [], {}
        week_starts = sorted(payload.keys())
        schedules = {}
        for week_start in week_starts:
            schedule = payload[week_start]
            archive_week(week_start, schedule)
            register_roles(schedule)
            schedules[week_start] = schedule
        if week_starts:
            save_last_schedule({"week_start": week_starts[-1], "schedule": schedules[week_starts[-1]]})
        return week_starts, schedules

    week_starts = sorted(w for w in week_starts_list if w)
    history = load_history()
    schedules = {w: history.get(w, {}) for w in week_starts}
    return week_starts, schedules


def export_file_stem(week_starts):
    first = datetime.strptime(week_starts[0], "%Y-%m-%d")
    last = datetime.strptime(week_starts[-1], "%Y-%m-%d")
    if first == last:
        return f"Schedule_{first.strftime('%Y_%m_%d')}"
    return f"Schedules_{first.strftime('%Y_%m_%d')}_to_{last.strftime('%Y_%m_%d')}"


@app.route("/export-multi-xlsx", methods=["POST"])
def export_multi_xlsx():
    employees = load_employees()
    week_starts, schedules = resolve_export_weeks(
        request.form.get("weeks_json"), request.form.getlist("week_starts")
    )
    if not week_starts:
        return "Select at least one week to export.", 400

    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()
    for week_start in week_starts:
        start_date, _end_date, date_range = week_date_range_text(week_start)
        rows = build_week_rows(employees, schedules.get(week_start, {}))
        title = safe_sheet_title(used_titles, start_date)
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        write_week_sheet(ws, date_range, rows)

    file_name = f"{export_file_stem(week_starts)}.xlsx"
    file_path = os.path.join(APP_DIR, file_name)
    wb.save(file_path)

    return send_file(file_path, as_attachment=True, download_name=file_name)


@app.route("/export-multi-pdf", methods=["POST"])
def export_multi_pdf():
    employees = load_employees()
    week_starts, schedules = resolve_export_weeks(
        request.form.get("weeks_json"), request.form.getlist("week_starts")
    )
    if not week_starts:
        return "Select at least one week to export.", 400

    elements = []
    for i, week_start in enumerate(week_starts):
        _start_date, _end_date, date_range = week_date_range_text(week_start)
        rows = build_week_rows(employees, schedules.get(week_start, {}))
        if i > 0:
            elements.append(PageBreak())
        elements.extend(build_pdf_week_elements(date_range, rows))

    file_name = f"{export_file_stem(week_starts)}.pdf"
    file_path = os.path.join(APP_DIR, file_name)

    doc = SimpleDocTemplate(
        file_path,
        pagesize=LETTER,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    doc.build(elements)

    return send_file(file_path, as_attachment=True, download_name=file_name)


# ------------------------------------
# SHIFT ROLES — canonical list, managed on its own (not per-person)
# ------------------------------------
@app.route("/roles/add", methods=["POST"])
def add_role():
    name = request.form.get("new_role", "").strip()
    if name:
        roles = load_roles()
        if name not in roles:
            save_roles(roles + [name])
            flash(f"Added role “{name}”.")
    return redirect(url_for("manage_employees"))


@app.route("/roles/delete", methods=["POST"])
def delete_role():
    name = request.form.get("role", "")
    roles = [r for r in load_roles() if r != name]
    save_roles(roles)
    return redirect(url_for("manage_employees"))


@app.route("/roles/cleanup", methods=["POST"])
def cleanup_roles():
    used = extract_roles_from_schedule(load_last_schedule().get("schedule", {}))
    for week_schedule in load_history().values():
        used |= extract_roles_from_schedule(week_schedule)

    roles = load_roles()
    kept = [r for r in roles if r in used]
    removed = len(roles) - len(kept)
    save_roles(kept)
    flash(f"Removed {removed} unused role(s)." if removed else "No unused roles found.")
    return redirect(url_for("manage_employees"))


# ------------------------------------
# EMPLOYEE MANAGEMENT PAGE
# ------------------------------------
@app.route("/employees", methods=["GET", "POST"])
def manage_employees():
    employees = load_employees()

    if request.method == "POST":
        # Add new employee
        new_emp = request.form.get("new_employee", "").strip()
        if new_emp and new_emp not in employees:
            employees.append(new_emp)

        # Delete selected employees — surgical: only these names are removed,
        # from the roster and from every saved/archived schedule. Everyone
        # else's records are independent, keyed by their own name, and are
        # never touched by this. A backup is snapshotted first so it can be
        # undone.
        to_delete = request.form.getlist("delete_emp")
        if to_delete:
            backup_employee_data()
            employees = [e for e in employees if e not in to_delete]
            remove_employees_from_schedules(to_delete)
            flash(f"Removed {len(to_delete)} team member(s). Use Undo to bring them back.")

        save_employees(employees)
        return redirect(url_for("manage_employees"))

    return render_template(
        "employees.html",
        employees=employees,
        roles=load_roles(),
        has_backup=os.path.exists(EMP_BACKUP_FILE),
    )


@app.route("/employees/undo", methods=["POST"])
def undo_employee_delete():
    if restore_employee_backup():
        flash("Restored the team and schedule data from before the last delete.")
    else:
        flash("No backup available to restore.")
    return redirect(url_for("manage_employees"))


if __name__ == "__main__":
    webbrowser.open("http://127.0.0.1:5000")
    app.run(debug=False)
